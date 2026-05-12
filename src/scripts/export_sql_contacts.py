import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ─────────────────────────────────────────────────────
SERVER   = r"SANTISIELE\SQLEXPRESS"
DATABASE = "SeniorHome"  

CONNECTION_STRING = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={SERVER};"
    f"DATABASE={DATABASE};"
    f"Trusted_Connection=yes;"
)
# SQL Server auth — uncomment if needed:
# CONNECTION_STRING = (
#     f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#     f"SERVER={SERVER};DATABASE={DATABASE};"
#     f"UID=your_user;PWD=your_password;"
# )

COLUMNS = [
    "nombres",
    "apellidos",
    "documento",
    "estado",
    "apellido_contacto",
    "nombre_contacto",
    "email_contacto",
    "documento_contacto",
    "cuit",
    "cuit_contacto",
]

OUTPUT_FILE = "sources/SeniorHome.xlsx"
# ───────────────────────────────────────────────────────────────


def main():
    cols = ", ".join(COLUMNS)
    query = f"""
        SELECT {cols}
        FROM contactos
        WHERE id <> 1
        ORDER BY id
    """

    with pyodbc.connect(CONNECTION_STRING) as conn:
        df = pd.read_sql(query, conn)

    print(f"Rows fetched: {len(df)}")

    df.to_excel(OUTPUT_FILE, index=False, sheet_name="Contacts")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font    = Font(name="Arial", size=10)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(cell.value or "")),
            max((len(str(ws.cell(r, col_idx).value or "")) for r in range(2, ws.max_row + 1)), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = cell_font

    wb.save(OUTPUT_FILE)
    print(f"File saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()