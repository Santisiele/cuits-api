import csv
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
import cuitonline

CSV_INPUT   = "sources/Poseidon.csv"
XLSX_OUTPUT = "sources/Filled-dbPoseidon.xlsx"
DELIMITER   = ";"
PAUSE_SEC   = 1.5

COLOR_DUPLICATES = "FFF2CC"
COLOR_ERROR      = "FCE4D6"
COLOR_NEW        = "E2EFDA"

def search_cuit(name):
    try:
        results = cuitonline.search(name.strip())
        if not results:
            return "ERROR - Not found", "error"
        if len(results) == 1:
            return results[0].cuit, "ok"
        all_cuits = " | ".join(f"{p.nombre} ({p.cuit})" for p in results)
        return f"DUPLICATES: {all_cuits}", "duplicates"
    except Exception as e:
        return f"ERROR - {e}", "error"

def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def main():
    with open(CSV_INPUT, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=DELIMITER)
        rows = list(reader)

    headers = list(rows[0].keys()) if rows else []
    print(f"✔ CSV loaded: {len(rows)} rows | Columns: {headers}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Result"

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = 22

    total      = len(rows)
    skipped    = 0
    processed  = 0
    errors     = 0
    duplicates = 0

    for row_idx, row in enumerate(rows, 2):
        existing_cuit = str(row.get("CUIT", "")).strip()

        if existing_cuit:
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
            skipped += 1
            continue

        name = row.get("Nombre completo", "").strip()
        print(f"[{row_idx-1}/{total}] Searching: {name} ...", end=" ")
        cuit_result, result_type = search_cuit(name)
        row["CUIT"] = cuit_result
        print(cuit_result)

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

        if result_type == "duplicates":
            fill = make_fill(COLOR_DUPLICATES)
            duplicates += 1
        elif result_type == "error":
            fill = make_fill(COLOR_ERROR)
            errors += 1
        else:
            fill = make_fill(COLOR_NEW)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

        processed += 1
        time.sleep(PAUSE_SEC)

    ws_summary = wb.create_sheet("Summary")
    summary = [
        ("Total rows",                total),
        ("Already had CUIT (skipped)", skipped),
        ("Queried",                   processed),
        ("Found (1 result)",          processed - duplicates - errors),
        ("Duplicates",                duplicates),
        ("Not found (ERROR)",         errors),
    ]
    for r, (label, val) in enumerate(summary, 1):
        ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=r, column=2, value=val)
    ws_summary.column_dimensions["A"].width = 30

    wb.save(XLSX_OUTPUT)
    print(f"\n✔ Done! Saved to: {XLSX_OUTPUT}")
    print(f"   Skipped: {skipped} | Found: {processed-duplicates-errors} | Duplicates: {duplicates} | Errors: {errors}")

if __name__ == "__main__":
    main()